from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from typing import List
import csv
import io
from backend.app.database import get_db
from backend.app.models import Tower, NetworkLink
from backend.app.schemas import Tower as TowerSchema, TowerCreate
from backend.app.schemas import NetworkLink as NetworkLinkSchema, NetworkLinkCreate
from backend.app.schemas import NetworkLink as NetworkLinkSchema, NetworkLinkCreate
from backend.app.services.cache import cache  # Cache para performance
from backend.app.dependencies import get_current_user

router = APIRouter(prefix="/towers", tags=["towers"])

@router.get("/", response_model=List[TowerSchema])
async def read_towers(skip: int = 0, limit: int = 100, db: AsyncSession = Depends(get_db)):
    # Tenta buscar do cache
    cache_key = f"towers_list_{skip}_{limit}"
    cached = await cache.get(cache_key)
    if cached:
        return cached
    
    # Se não está no cache, busca do banco
    result = await db.execute(select(Tower).offset(skip).limit(limit))
    towers = result.scalars().all()
    
    # Salva no cache por 30 segundos
    await cache.set(cache_key, towers, ttl_seconds=30)
    
    return towers

@router.post("/", response_model=TowerSchema)
async def create_tower(tower: TowerCreate, db: AsyncSession = Depends(get_db), current_user=Depends(get_current_user)):
    db_tower = Tower(**tower.model_dump())
    db.add(db_tower)
    await db.commit()
    await db.refresh(db_tower)
    
    # Invalida o cache de torres
    await cache.delete("towers_list_0_100")
    
    return db_tower

@router.post("/import_csv")
async def import_towers_csv_route(file: UploadFile = File(...), db: AsyncSession = Depends(get_db), current_user=Depends(get_current_user)):
    """Importa torres via arquivo CSV ou Excel (.xlsx)"""
    try:
        content = await file.read()
        filename = file.filename.lower()
        rows = []
        columns = []

        # Lógica para EXCEL (XLSX)
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            try:
                from openpyxl import load_workbook
                wb = load_workbook(filename=io.BytesIO(content), data_only=True)
                ws = wb.active # Pega a primeira aba
                
                # Ler todas as linhas
                all_rows = list(ws.iter_rows(values_only=True))
                if not all_rows:
                     raise HTTPException(status_code=400, detail="Planilha vazia")
                
                # Cabeçalho é a primeira linha
                columns = [str(c).lower().strip() if c else "" for c in all_rows[0]]
                
                # Dados são o resto
                for row_data in all_rows[1:]:
                    row_dict = {}
                    for idx, val in enumerate(row_data):
                        if idx < len(columns) and columns[idx]:
                            row_dict[columns[idx]] = val
                    if any(row_dict.values()): # Se a linha não for vazia
                        rows.append(row_dict)
                        
            except ImportError:
                 raise HTTPException(status_code=500, detail="Biblioteca openpyxl não instalada no servidor.")
            except Exception as e:
                 raise HTTPException(status_code=400, detail=f"Erro ao ler Excel: {str(e)}")

        # Lógica para CSV (Texto)
        else:
            # Decode UTF-8 e remove BOM se existir
            try:
                decoded = content.decode('utf-8-sig')
            except UnicodeDecodeError:
                # Tenta latin-1 se utf-8 falhar (comum no Brasil)
                decoded = content.decode('latin-1')

            # Tenta detectar dialeto
            try:
                dialect = csv.Sniffer().sniff(decoded[:2048], delimiters=',;\t')
            except:
                class Dialect(csv.Dialect):
                    delimiter = '\t'
                    quotechar = '"'
                    doublequote = True
                    skipinitialspace = True
                    lineterminator = '\n'
                    quoting = csv.QUOTE_MINIMAL
                dialect = Dialect

            f = io.StringIO(decoded)
            reader = csv.DictReader(f, dialect=dialect)
            columns = reader.fieldnames or []
            rows = list(reader)
        
        # --- Processamento Unificado (Normalização de Colunas) ---
        
        # Mapeamento de colunas (Excel ou CSV)
        # Procura índice ou chave das colunas
        def find_key(candidates, keys):
            for k in keys:
                if k in candidates: return k
                # Tenta match parcial ex: "Nome da Torre" -> "nome"
                for cand in candidates:
                    if k in cand: return cand
            return None

        # Normaliza chaves de cada linha
        normalized_rows = []
        
        # Identificar nomes reais das colunas no arquivo
        keys_map = {} # map internal_name -> file_column_name
        
        # Lista de candidatos
        possible_cols = list(rows[0].keys()) if rows else columns
        # Se for CSV list(reader) já são dicts, se excel rows são dicts. 
        # Mas excel columns headers podem estar sujos.
        
        # Função auxiliar para pegar valor flexível
        def get_val(row, aliases):
            for alias in aliases:
                for key in row.keys():
                    if alias in key.lower():
                        return row[key]
            return None

        if not rows:
             raise HTTPException(status_code=400, detail="Arquivo vazio ou sem dados")

        imported_count = 0
        skipped_count = 0
        
        for row in rows:
            # Busca nome
            name = get_val(row, ['nome', 'name', 'torre'])
            if not name or not str(name).strip(): continue
            name = str(name).strip()
            
            # Checar duplicidade
            stmt = select(Tower).where(Tower.name == name)
            res = await db.execute(stmt)
            if res.scalars().first():
                skipped_count += 1
                continue

            lat = get_val(row, ['lat', 'latitude'])
            lon = get_val(row, ['lon', 'long', 'longitude'])
            obs = get_val(row, ['obs', 'descri', 'description'])
            ip = get_val(row, ['ip', 'address'])
            
            # Limpar Floats
            def clean_float(v):
                if not v: return None
                s = str(v).strip().replace(',', '.')
                try: return float(s)
                except: return None
            
            new_tower = Tower(
                name=name,
                latitude=clean_float(lat),
                longitude=clean_float(lon),
                ip=str(ip).strip() if ip else None,
                observations=str(obs) if obs else None,
                is_online=True
            )
            db.add(new_tower)
            imported_count += 1
        
        await db.commit()
        
        # Invalida cache
        await cache.delete("towers_list_0_100")
        
        return {"message": "Importação concluída", "imported": imported_count, "skipped": skipped_count}

    except Exception as e:
        await db.rollback()
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Erro na importação: {str(e)}")

# --- Network Links Endpoints (MUST come before /{tower_id} to avoid route conflict) ---

@router.get("/links", response_model=List[NetworkLinkSchema])
async def read_links(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(NetworkLink))
    return result.scalars().all()

@router.post("/links", response_model=NetworkLinkSchema)
async def create_link(link: NetworkLinkCreate, db: AsyncSession = Depends(get_db), current_user=Depends(get_current_user)):
    db_link = NetworkLink(**link.model_dump())
    db.add(db_link)
    await db.commit()
    await db.refresh(db_link)
    return db_link

@router.delete("/links/{link_id}")
async def delete_link(link_id: int, db: AsyncSession = Depends(get_db), current_user=Depends(get_current_user)):
    db_link = await db.get(NetworkLink, link_id)
    if not db_link:
        raise HTTPException(status_code=404, detail="Link not found")
    await db.delete(db_link)
    await db.commit()
    return {"message": "Link deleted"}

# --- Tower-specific endpoints (generic /{tower_id} must come AFTER specific routes) ---

@router.get("/{tower_id}", response_model=TowerSchema)
async def read_tower(tower_id: int, db: AsyncSession = Depends(get_db)):
    db_tower = await db.get(Tower, tower_id)
    if db_tower is None:
        raise HTTPException(status_code=404, detail="Tower not found")
    return db_tower

@router.delete("/{tower_id}")
async def delete_tower(tower_id: int, db: AsyncSession = Depends(get_db), current_user=Depends(get_current_user)):
    db_tower = await db.get(Tower, tower_id)
    if not db_tower:
        raise HTTPException(status_code=404, detail="Tower not found")
    await db.delete(db_tower)
    await db.commit()
    
    # Invalida o cache de torres
    await cache.delete("towers_list_0_100")
    
    return {"message": "Tower deleted"}
